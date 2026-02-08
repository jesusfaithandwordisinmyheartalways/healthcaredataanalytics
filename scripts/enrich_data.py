from openpyxl import load_workbook, Workbook
import random
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
input_path = os.path.join(BASE_DIR, 'data', 'healthcare_data_cleaned.xlsx')
output_path = os.path.join(BASE_DIR, 'data', 'healthcare_data_enriched.xlsx')

wb = load_workbook(input_path)
ws = wb.active

headers = [cell.value for cell in ws[1]]

# NEW HEADERS (do not remove originals)
new_headers = headers + [
    "Year",
    "Month",
    "Region",
    "State",
    "ALL_Visits",
    "Billing_Method",
    "Yearly_Billing_Amount",
    "Total_Billing_2021_2025",
    "Insurance_Total_Sales",
    "Department_Total_Sales",
    "Monthly_Sales",
    "Total_Company_Billing_By_Year"
]

wb_new = Workbook()
ws_new = wb_new.active
ws_new.append(new_headers)

years = [2021, 2022, 2023, 2024, 2025]
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
regions = ["All","Midwest","Northeast","South","Southeast","West"]
states = [
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN","IA",
    "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ",
    "NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT",
    "VA","WA","WV","WI","WY"
]
billing_methods = [
    "Medicare",
    "Medicaid",
    "Military/Veterans Benefits",
    "Self Pay - Credit/Debit",
    "Assistance Programs"
]

company_year_totals = {year: random.randint(20000000, 60000000) for year in years}

for row in ws.iter_rows(min_row=2):
    base_data = [cell.value for cell in row]

    year = random.choice(years)
    month = random.choice(months)
    region = random.choice(regions)
    state = random.choice(states)
    billing_method = random.choice(billing_methods)

    yearly_billing = random.randint(5000, 50000)
    total_5yr = sum(random.randint(30000, 90000) for _ in years)

    insurance_total = random.randint(1000000, 20000000)
    department_total = random.randint(500000, 15000000)
    monthly_sales = random.randint(0, 60000000)

    ws_new.append(
        base_data + [
            year,
            month,
            region,
            state,
            "ALL",
            billing_method,
            yearly_billing,
            total_5yr,
            insurance_total,
            department_total,
            monthly_sales,
            company_year_totals[year]
        ]
    )

wb_new.save(output_path)
print("Enriched healthcare dataset saved to:", output_path)