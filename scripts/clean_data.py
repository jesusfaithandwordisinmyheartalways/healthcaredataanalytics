from openpyxl import load_workbook, Workbook
import os

# Build absolute path to Excel file
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(BASE_DIR, 'data', 'healthcare_data.xlsx')

print("Loading file from:", file_path)

# Load workbook
wb = load_workbook(file_path)
sheet = wb.active

# Remove rows with missing Billing_Amount
rows_to_keep = []
for row in sheet.iter_rows(min_row=2):
    if row[6].value is not None:
        rows_to_keep.append([cell.value for cell in row])

# Save cleaned data
wb_new = Workbook()
ws = wb_new.active

# Add header row
ws.append([cell.value for cell in sheet[1]])

# Add filtered rows
for row in rows_to_keep:
    ws.append(row)

# Save cleaned file
cleaned_path = os.path.join(BASE_DIR, 'data', 'healthcare_data_cleaned.xlsx')
wb_new.save(cleaned_path)

print("Cleaned file saved:", cleaned_path)